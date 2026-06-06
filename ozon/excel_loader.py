from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
import pandas as pd


REQUIRED_COLUMNS = [
    "offer_id",
    "name_ru",
    "description_ru",
    "description_category_id",
    "type_id",
    "price",
    "currency_code",
    "vat",
    "stock",
    "weight",
    "depth",
    "width",
    "height",
    "primary_image",
    "images",
    "attributes_json",
]

OPTIONAL_COLUMNS = [
    "old_price",
    "barcode",
    "color",
    "size",
    "material",
    "brand",
    "model",
    "unit_product_quantity",
    "items_in_product",
    "category_name",
    "note",
]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

HEADER_ALIASES = {
    "offer_id": ["offer_id", "货号*", "货号", "商品编号", "商品货号"],
    "name_ru": ["name_ru", "商品名称", "商品名称*", "俄语商品标题", "名称"],
    "description_ru": ["description_ru", "简介", "商品描述", "俄语商品描述"],
    "description_category_id": ["description_category_id", "Ozon 类目 ID", "类目 ID", "description category id"],
    "type_id": ["type_id", "商品类型 ID", "类型 ID", "type id"],
    "price": ["price", "价格，CNY*", "价格", "售价", "售价，RUB"],
    "old_price": ["old_price", "折扣前价格，CNY", "划线价"],
    "currency_code": ["currency_code", "币种", "货币"],
    "vat": ["vat", "VAT", "增值税"],
    "stock": ["stock", "库存", "库存数量"],
    "weight": ["weight", "毛重，克*", "毛重，克", "商品重量，克", "重量"],
    "depth": ["depth", "包装长度，毫米*", "包装长度，毫米", "长，毫米", "长度"],
    "width": ["width", "包装宽度，毫米*", "包装宽度，毫米", "宽，毫米", "宽度"],
    "height": ["height", "包装高度，毫米*", "包装高度，毫米", "高，毫米", "高度"],
    "primary_image": ["primary_image", "主图链接*", "主图链接", "主图", "主图文件名"],
    "images": ["images", "附加图片链接", "附图", "附图文件名", "图片"],
    "attributes_json": ["attributes_json", "JSON对应内容", "JSON 对应内容", "属性 JSON", "Ozon 属性 JSON"],
    "barcode": ["barcode", "条形码（序列号/EAN）", "条形码", "EAN"],
    "color": ["color", "商品颜色", "颜色"],
    "size": ["size", "动物服装/配件的尺寸", "尺码", "尺寸"],
    "material": ["material", "材料", "材质"],
    "brand": ["brand", "品牌*", "品牌"],
    "model": ["model", "型号名称（针对合并为一张商品卡片）*", "型号名称", "型号"],
    "unit_product_quantity": ["unit_product_quantity", "统一计量单位中的商品数量", "统一计量单位商品数量"],
    "items_in_product": ["items_in_product", "一个商品中的件数", "商品中的件数", "件数"],
    "category_name": ["category_name", "类目名称"],
    "note": ["note", "备注", "人工审核备注"],
}

DEFAULT_VALUES = {
    "currency_code": "RUB",
    "vat": 0,
    "stock": 0,
    "attributes_json": "{}",
}

TEMPLATE_ROW_MARKERS = {"必填字段", "选填字段", "字段", "示例", "说明"}


def _clean_cell(value):
    if pd.isna(value):
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _norm_header(value) -> str:
    return str(value or "").replace("\n", "").replace(" ", "").strip().lower()


def _rename_with_aliases(df: pd.DataFrame) -> pd.DataFrame:
    alias_to_field = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_to_field[_norm_header(alias)] = field

    rename_map = {}
    used_fields = set()
    for col in df.columns:
        normalized = _norm_header(col)
        target = alias_to_field.get(normalized)
        if target and target not in used_fields:
            rename_map[col] = target
            used_fields.add(target)
    return df.rename(columns=rename_map)


def _read_best_sheet(path: Path) -> pd.DataFrame:
    best_df = None
    best_score = -1
    # Ozon 中文模板通常第 2 行才是真正字段名，所以这里尝试多个表头行。
    for header_row in [0, 1, 2, 3, 4]:
        try:
            df = pd.read_excel(path, dtype=object, engine="openpyxl", header=header_row)
        except Exception:
            continue
        df.columns = [str(col).strip() for col in df.columns]
        mapped = _rename_with_aliases(df)
        score = len(set(mapped.columns) & set(ALL_COLUMNS))
        if score > best_score:
            best_score = score
            best_df = mapped
    if best_df is None:
        best_df = pd.read_excel(path, dtype=object, engine="openpyxl")
        best_df.columns = [str(col).strip() for col in best_df.columns]
    return best_df


def load_products_from_excel(path: str | Path) -> Dict:
    excel_path = Path(path)
    df = _read_best_sheet(excel_path)

    for col, default_value in DEFAULT_VALUES.items():
        if col not in df.columns:
            df[col] = default_value
        else:
            df[col] = df[col].apply(lambda v: default_value if pd.isna(v) or str(v).strip() == "" else v)

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    for col in ALL_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[ALL_COLUMNS]
    df = df.map(_clean_cell)
    df = df.dropna(how="all")
    df = df[~df["offer_id"].astype(str).str.strip().isin(TEMPLATE_ROW_MARKERS)]
    offer_text = df["offer_id"].astype(str).str.strip()
    name_text = df["name_ru"].astype(str).str.strip()
    df = df[
        ~(
            offer_text.str.contains("请输入商品货号", na=False)
            | offer_text.str.contains("docs.ozon", na=False)
            | name_text.str.contains("名称要求", na=False)
        )
    ]
    df = df[
        ~(
            (df["offer_id"].astype(str).str.strip() == "")
            & (df["name_ru"].astype(str).str.strip() == "")
            & (df["primary_image"].astype(str).str.strip() == "")
        )
    ]

    products: List[Dict] = df.to_dict(orient="records")
    return {
        "path": str(excel_path),
        "products": products,
        "missing_columns": missing_columns,
        "row_count": len(products),
        "columns": list(df.columns),
    }


def create_sample_excel(path: str | Path) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    sample = pd.DataFrame(
        [
            {
                "offer_id": "TEST-001",
                "name_ru": "Дождевик для собак",
                "description_ru": "Легкий дождевик для прогулок в дождливую погоду.",
                "description_category_id": "",
                "type_id": "",
                "price": 999,
                "old_price": 1299,
                "currency_code": "RUB",
                "vat": 0,
                "stock": 10,
                "weight": 150,
                "depth": 300,
                "width": 220,
                "height": 30,
                "primary_image": "TEST-001_main.jpg",
                "images": "TEST-001_1.jpg,TEST-001_2.jpg",
                "attributes_json": "{}",
                "brand": "Нет бренда",
                "unit_product_quantity": 1,
                "items_in_product": 1,
                "color": "",
                "size": "",
                "material": "",
                "model": "",
                "category_name": "",
                "note": "示例行，请改成真实商品资料",
            }
        ]
    )
    sample.to_excel(output, index=False)
    return output


OZON_TEMPLATE_ALIASES = {
    "offer_id": ["货号*", "货号", "offer_id"],
    "name_ru": ["商品名称", "商品名称*", "name_ru"],
    "description_ru": ["简介", "商品描述", "description_ru"],
    "description_category_id": ["description_category_id", "Ozon 类目 ID", "类目 ID"],
    "type_id": ["type_id", "商品类型 ID", "类型 ID"],
    "price": ["价格，CNY*", "价格", "售价", "price"],
    "old_price": ["折扣前价格，CNY", "折扣前价格", "划线价", "old_price"],
    "stock": ["stock", "库存", "库存数量"],
    "weight": ["毛重，克*", "毛重，克", "商品重量，克", "weight"],
    "product_weight": ["商品克重", "商品克重，克", "商品克重 g", "product_weight"],
    "depth": ["包装长度，毫米*", "包装长度，毫米", "depth"],
    "width": ["包装宽度，毫米*", "包装宽度，毫米", "width"],
    "height": ["包装高度，毫米*", "包装高度，毫米", "height"],
    "primary_image": ["主图链接*", "主图链接", "primary_image"],
    "images": ["附加图片链接", "附加图片", "images"],
    "brand": ["品牌*", "品牌", "brand"],
    "model": ["型号名称（针对合并为一张商品卡片）*", "型号名称", "model"],
    "unit_product_quantity": ["统一计量单位中的商品数量", "统一计量单位商品数量"],
    "items_in_product": ["一个商品中的件数", "商品中的件数", "件数"],
    "animal_gender": ["动物的性别"],
    "color": ["商品颜色", "颜色"],
    "color_name_ru": ["颜色名称"],
    "type": ["类型*", "类型"],
    "target_pet": ["专为*", "专为"],
    "clothing_type": ["服装类型"],
    "hashtags": ["#主题标签", "主题标签"],
    "pet_size": ["动物服装/配件的尺寸", "宠物尺寸"],
    "neck_max_cm": ["最大颈围，厘米", "最大颈围", "颈围最大值"],
    "neck_min_cm": ["最小颈围，厘米", "最小颈围", "颈围最小值"],
    "chest_min_cm": ["最小胸围，厘米", "最小胸围", "胸围最小值"],
    "chest_max_cm": ["最大胸围，厘米", "最大胸围", "胸围最大值"],
    "back_length_cm": ["背部长度，厘米", "背长，厘米", "背长"],
    "material": ["材料", "材质"],
    "brand_country": ["品牌国家"],
    "origin_country": ["原产国"],
    "package_size_cm": ["包装尺寸（长X宽x高），厘米", "包装尺寸（长X宽x高）,厘米"],
}

API_HELPER_TEMPLATE_COLUMNS = {
    "description_category_id": "description_category_id",
    "type_id": "type_id",
    "old_price": "old_price",
    "product_weight": "product_weight",
    "stock": "stock",
}


def _normalize_template_header(value) -> str:
    return str(value or "").replace("\n", "").replace(" ", "").strip().lower()


def _find_template_header_row(sheet) -> int:
    wanted = {"货号*", "商品名称", "主图链接*"}
    best_row = 1
    best_score = -1
    max_row = min(sheet.max_row, 12)
    for row_idx in range(1, max_row + 1):
        values = {_normalize_template_header(cell.value) for cell in sheet[row_idx]}
        score = sum(1 for item in wanted if _normalize_template_header(item) in values)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def _build_template_column_map(sheet, header_row: int) -> Dict[str, int]:
    alias_to_field = {}
    for field, aliases in OZON_TEMPLATE_ALIASES.items():
        for alias in aliases:
            alias_to_field[_normalize_template_header(alias)] = field

    columns: Dict[str, int] = {}
    for cell in sheet[header_row]:
        field = alias_to_field.get(_normalize_template_header(cell.value))
        if field and field not in columns:
            columns[field] = cell.column
    return columns


def _ensure_api_helper_columns(sheet, header_row: int, columns: Dict[str, int]) -> None:
    for field, header in API_HELPER_TEMPLATE_COLUMNS.items():
        if columns.get(field):
            continue
        col_idx = sheet.max_column + 1
        sheet.cell(header_row, col_idx, value=header)
        columns[field] = col_idx


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col_idx)
        target = sheet.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _mm_to_cm(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text) / 10
    except ValueError:
        return None


def _format_package_size_cm(depth, width, height) -> str:
    values = [_mm_to_cm(item) for item in [depth, width, height]]
    if any(item is None for item in values):
        return ""
    return "x".join(f"{item:g}" for item in values if item is not None)


def fill_ozon_template(template_path: str | Path, products: List[Dict], output_path: str | Path) -> Path:
    template = Path(template_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    sheet = workbook.active
    header_row = _find_template_header_row(sheet)
    columns = _build_template_column_map(sheet, header_row)
    _ensure_api_helper_columns(sheet, header_row, columns)

    start_row = max(header_row + 1, 7)
    if sheet.max_row < start_row:
        sheet.insert_rows(start_row)

    for offset, product in enumerate(products):
        row_idx = start_row + offset
        if row_idx > sheet.max_row:
            sheet.insert_rows(row_idx)
        _copy_row_style(sheet, start_row, row_idx)

        tags = product.get("tags_ru") or []
        if isinstance(tags, str):
            tags_text = tags
        else:
            tags_text = " ".join(str(item).strip() for item in tags if str(item).strip())

        additional_images = product.get("additional_images") or []
        if isinstance(additional_images, str):
            images_text = additional_images
        else:
            images_text = ",".join(str(item).strip() for item in additional_images if str(item).strip())

        depth = product.get("depth", "")
        width = product.get("width", "")
        height = product.get("height", "")
        package_size = _format_package_size_cm(depth, width, height)

        values = {
            "offer_id": product.get("offer_id", ""),
            "name_ru": product.get("name_ru", ""),
            "description_ru": product.get("description_ru", ""),
            "description_category_id": product.get("description_category_id", ""),
            "type_id": product.get("type_id", ""),
            "price": product.get("price", ""),
            "old_price": product.get("old_price", ""),
            "stock": product.get("stock", ""),
            "weight": product.get("weight", ""),
            "product_weight": product.get("product_weight", ""),
            "depth": depth,
            "width": width,
            "height": height,
            "primary_image": product.get("primary_image", ""),
            "images": images_text,
            "brand": product.get("brand") or "无品牌",
            "model": product.get("model", ""),
            "unit_product_quantity": product.get("unit_product_quantity", 1) or 1,
            "items_in_product": product.get("items_in_product", 1) or 1,
            "animal_gender": "男女两用的",
            "color": product.get("color", ""),
            "color_name_ru": product.get("color_name_ru", ""),
            "type": product.get("type") or "宠物服装",
            "target_pet": product.get("target_pet") or "对于狗",
            "clothing_type": product.get("clothing_type") or product.get("product_type_ru", ""),
            "hashtags": tags_text,
            "pet_size": product.get("pet_size", ""),
            "neck_max_cm": product.get("neck_max_cm", ""),
            "neck_min_cm": product.get("neck_min_cm", ""),
            "chest_min_cm": product.get("chest_min_cm", ""),
            "chest_max_cm": product.get("chest_max_cm", ""),
            "back_length_cm": product.get("back_length_cm", ""),
            "material": product.get("material", ""),
            "brand_country": "中国",
            "origin_country": "中国",
            "package_size_cm": package_size,
        }

        for field, value in values.items():
            col_idx = columns.get(field)
            if col_idx:
                sheet.cell(row_idx, col_idx, value=value)

    workbook.save(output)
    return output
